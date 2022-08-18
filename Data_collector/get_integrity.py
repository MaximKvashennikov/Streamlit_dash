import pandas as pd
from Data_collector.number_week import three_week
from streamlit import cache
import zipfile
import rarfile
import patoolib
import os
from openpyxl.reader.excel import load_workbook


def get_path(path_user="/mnt/rrl_integrity_share"):
    """ Получение последнего файла с пролетами из папки """

    list_files_sw = [s for s in os.listdir(path_user)
                     if os.path.isfile(os.path.join(path_user, s))]
    list_files_sw.sort(key=lambda s: os.path.getmtime(os.path.join(path_user, s)))
    path_file = path_user + "//" + list_files_sw[-1]

    return path_file


def check_format():
    """ Проверка формата архива """
    return get_path().split(".")[-1]


def get_last_week():
    """ Получение номера недели из последнего существующего файла """

    return (get_path().replace("_new", "").split("_")[-1]).split(".")[0]


def unpack_handler():
    if check_format() == "zip":
        return unpack_file_zip()
    elif check_format() == "rar":
        # return unpack_file_rar()
        return unpack_file_any()
    else:
        return unpack_file_any()


def unpack_file_rar():
    """ Распаковка .rar """

    # rarfile.UNRAR_TOOL = "D:\Python\Streamlit\Data_collector\\UnRAR.exe"
    path = get_path()
    dir_out = "/mnt/cpfolders_share/Operation Group/Отчеты RRL Capacity_4pica/Weekly_RRL_Integrity"
    with rarfile.RarFile(path, "r") as rf:
        rf.extractall(dir_out)
    return f"{dir_out}//{path.split('//')[-1].split('.')[0]}.xlsx".replace('\\', '//')


def unpack_file_any():
    """ Распаковка любого архива """

    path = get_path()
    dir_out = "/mnt/cpfolders_share/Operation Group/Отчеты RRL Capacity_4pica/Weekly_RRL_Integrity"
    patoolib.extract_archive(path,
                             outdir=dir_out,
                             interactive=False
                             )
    return f"{get_path(path_user=dir_out)}".replace('\\', '//')


def unpack_file_zip():
    """ Распаковка .zip """

    path = get_path()
    archive = zipfile.ZipFile(path, 'r')
    return archive.open(path.split('//')[-1].replace('zip', 'xlsx'))


def get_annual_integrity():
    path_integrity = "/mnt/cpfolders_share/Operation Group/RRL_integrity_Ежегодный свод.xlsx"

    xls = pd.ExcelFile(path_integrity)
    df = pd.read_excel(
        path_integrity,
        sheet_name=xls.sheet_names[-1],
    )

    return df


def get_week_integrity():
    df = pd.read_excel(
        unpack_handler(),
        sheet_name="Weekly_Data",
    )

    return df


@cache(allow_output_mutation=True)
def get_percentage_of_overloaded(df):
    week_previous_integrity = f"w{get_last_week()[-2:]}"
    df = df[["Label", week_previous_integrity]].head(2)
    if type(df.iloc[0, 1]) == str:
        df.iloc[0, 1] = f"{round(float(df.iloc[0, 1][:-1].replace(',', '.')), 2)}%".replace('.', ',')
    elif df.iloc[0, 1].dtype == 'float64':
        df.iloc[0, 1] = f"{round(float(df.iloc[0, 1]) * 100, 2)}%".replace('.', ',')
    df.iloc[1, 1] = f"{round(float(df.iloc[1, 1]))}"
    df = df.style.set_properties(**{'background-color': '#E6E6FA',
                                    'font-size': '14pt',
                                    })
    return df


@cache(allow_output_mutation=True)
def set_annual_integrity(df):
    path_integrity = "/mnt/cpfolders_share/Operation Group/RRL_integrity_Ежегодный свод.xlsx"
    last_week = get_last_week()
    df = df[(df["Region"] == "RUSSIA") & (df["Label"] != "Total RRL")]
    df = df[f"{last_week}"]
    percent = f"{round(df.iloc[0] * 100, 2)}%".replace('.', ',')
    quantity = f"{round(df.iloc[1])}"

    wb = load_workbook(path_integrity)
    sheet = wb.worksheets[-1]
    column_week = ([cell for cell in list(sheet.rows)[0] if cell.value == f"w{last_week[-2:]}"])[0].column
    sheet.cell(row=2, column=column_week).value = percent
    sheet.cell(row=3, column=column_week).value = quantity
    wb.save(path_integrity)

    return percent, quantity


if __name__ == '__main__':
    # print(set_annual_integrity(get_week_integrity()))
    # print(get_last_week())
    print(unpack_file_any())
